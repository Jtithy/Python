#Author: Tithy
#Date: 2026-07-20
#Description: Working with Excel files in Python.

import openpyxl as xl  # type: ignore[import-not-found]
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
print(cell.value)
print(sheet.max_row)

for row in range(1, sheet.max_row + 1):
    print(row)